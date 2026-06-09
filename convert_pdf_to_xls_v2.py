import sys
import os

try:
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("正在安装所需依赖...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pdfplumber", "openpyxl"])
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

print("重新转换医保培训PDF，优化表格提取...")

pdf_path = '/Users/yantu/Desktop/医保培训.pdf'
output_path = '/Users/yantu/Desktop/医保培训.xlsx'

if not os.path.exists(pdf_path):
    print(f"错误：找不到文件 {pdf_path}")
    sys.exit(1)

pdf = pdfplumber.open(pdf_path)
wb = Workbook()

print(f"PDF共有 {len(pdf.pages)} 页")

for page_num, page in enumerate(pdf.pages, 1):
    print(f"正在处理第 {page_num} 页...")
    
    ws = wb.create_sheet(title=f"第{page_num}页")
    
    # 尝试多种方法提取表格
    tables = page.extract_tables(table_settings={
        "vertical_strategy": "text", 
        "horizontal_strategy": "text",
        "edge_min_length": 3
    })
    
    if tables:
        print(f"  找到 {len(tables)} 个表格")
        row = 1
        for table in tables:
            for table_row in table:
                for col_idx, cell in enumerate(table_row, 1):
                    if cell:
                        cleaned = str(cell).strip()
                        if cleaned:
                            ws.cell(row=row, column=col_idx, value=cleaned)
                row += 1
    else:
        # 如果没找到表格，尝试提取文本
        text = page.extract_text()
        if text:
            lines = text.split('\n')
            for row_idx, line in enumerate(lines, 1):
                if line.strip():
                    ws.cell(row=row_idx, column=1, value=line.strip())
                    ws.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True, vertical='center')

# 删除默认创建的空工作表
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

pdf.close()

# 调整所有工作表的列宽
for ws in wb.worksheets:
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column if cell.value]
        if column:
            max_length = max(len(str(cell.value)) for cell in column) + 2
            ws.column_dimensions[column[0].column_letter].width = min(max_length, 50)

wb.save(output_path)
print(f"\n✅ 转换完成！")
print(f"输出文件: {output_path}")
print(f"处理了 {len(pdf.pages)} 页，共创建了 {len(wb.worksheets)} 个工作表")