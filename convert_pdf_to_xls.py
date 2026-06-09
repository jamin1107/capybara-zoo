import sys
import os

try:
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("正在安装所需依赖...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pdfplumber", "openpyxl"])
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

print("开始转换医保培训PDF...")

pdf_path = '/Users/yantu/Desktop/医保培训.pdf'
output_path = '/Users/yantu/Desktop/医保培训.xls'

if not os.path.exists(pdf_path):
    print(f"错误：找不到文件 {pdf_path}")
    sys.exit(1)

# 打开PDF文件
pdf = pdfplumber.open(pdf_path)

# 创建Excel工作簿
wb = Workbook()
ws = wb.active
ws.title = "医保培训"

print(f"PDF共有 {len(pdf.pages)} 页")

current_row = 1
tables_extracted = 0

for page_num, page in enumerate(pdf.pages, 1):
    print(f"正在处理第 {page_num} 页...")
    
    # 提取文本内容
    text = page.extract_text()
    if text:
        lines = text.split('\n')
        for line in lines:
            if line.strip():
                ws.cell(row=current_row, column=1, value=line)
                ws.cell(row=current_row, column=1).alignment = Alignment(vertical='center', wrap_text=True)
                current_row += 1
    
    # 提取表格
    tables = page.extract_tables()
    for table in tables:
        tables_extracted += 1
        for row_data in table:
            for col_idx, cell_value in enumerate(row_data, 1):
                if cell_value:
                    ws.cell(row=current_row, column=col_idx, value=str(cell_value))
                    ws.cell(row=current_row, column=col_idx).alignment = Alignment(vertical='center')
            current_row += 1

pdf.close()

# 调整列宽
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column if cell.value]
    if column:
        max_length = max(len(str(cell.value)) for cell in column) + 2
        ws.column_dimensions[column[0].column_letter].width = min(max_length, 50)

wb.save(output_path)
print(f"\n✅ 转换完成！")
print(f"输出文件: {output_path}")
print(f"处理了 {len(pdf.pages)} 页，提取了 {tables_extracted} 个表格")