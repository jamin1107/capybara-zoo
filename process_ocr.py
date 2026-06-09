import sys
import os

print("正在安装OCR所需依赖...")

try:
    import pdf2image
    from PIL import Image
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("正在安装Python依赖...")
    import subprocess
    packages = ['pdf2image', 'pillow', 'openpyxl']
    for pkg in packages:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])
        except:
            pass

pdf_path = '/Users/yantu/Desktop/医保培训.pdf'
output_path = '/Users/yantu/Desktop/医保培训.xlsx'

print(f"开始处理PDF: {pdf_path}")

from pdf2image import convert_from_path

try:
    # 先尝试转换前5页看看
    print("正在将PDF转换为图片...")
    pages = convert_from_path(pdf_path, dpi=150, first_page=1, last_page=5)
    print(f"✅ 成功转换 {len(pages)} 页为图片")
    
except Exception as e:
    print(f"\n❌ PDF转换失败: {e}")
    print("\n需要先安装 poppler 工具")
    print("\n请在终端运行:")
    print("  brew install poppler")
    print("\n如果没有Homebrew，请访问 https://brew.sh/ 安装")
    
    # 提供一个替代方案说明
    print("\n\n========== 替代方案 ==========")
    print("\n1. 访问 https://smallpdf.com/pdf-to-excel")
    print("2. 上传PDF文件")
    print("3. 免费下载转换后的Excel")
    print("\n或使用其他在线OCR工具")
    sys.exit(1)

print("\n现在开始OCR识别...")

# 创建Excel
wb = Workbook()

from PIL import Image

print(f"准备好处理 {len(pages)} 页")

# 由于没有安装Tesseract OCR引擎，我们先将PDF转成图片并保存
images_dir = '/Users/yantu/Desktop/test/pdf_images'
os.makedirs(images_dir, exist_ok=True)

for i, image in enumerate(pages, 1):
    image_path = f'{images_dir}/page_{i}.png'
    image.save(image_path)
    print(f"已保存第{i}页图片: {image_path}")

print(f"\n✅ PDF转换为图片成功！")
print(f"图片保存在: {images_dir}/")
print("\n接下来需要OCR软件来识别这些图片...")

print("\n\n========== 下一步操作 ==========")
print("\n由于缺少OCR引擎，请按以下步骤操作:")
print("\n方法1: 安装Tesseract")
print("  brew install tesseract")
print("  然后我可以继续处理")
print("\n方法2: 使用在线OCR")
print("  访问 https://www.ilovepdf.com/ocr")
print("  上传PDF，选择中文识别")
print("\n方法3: 用WPS/Microsoft Office")
print("  直接用WPS打开PDF，导出为Excel")
print("\n\n我已将PDF前5页转为图片，保存在桌面上的test/pdf_images文件夹中")

# 创建一个说明文档
ws = wb.active
ws.title = "说明"
ws.cell(row=1, column=1, value="PDF转Excel说明")
ws.cell(row=2, column=1, value="由于PDF是扫描件，需要OCR识别")
ws.cell(row=3, column=1, value="PDF图片已保存到: " + images_dir)
ws.cell(row=5, column=1, value="推荐方案：")
ws.cell(row=6, column=1, value="1. 用 WPS Office 打开PDF，直接导出Excel")
ws.cell(row=7, column=1, value="2. 访问 https://smallpdf.com/pdf-to-excel")
ws.cell(row=8, column=1, value="3. 或者 brew install tesseract 后我继续处理")

ws.column_dimensions['A'].width = 60

wb.save(output_path.replace('.xlsx', '_说明.xlsx'))

print(f"\n已保存说明文件: {output_path.replace('.xlsx', '_说明.xlsx')}")