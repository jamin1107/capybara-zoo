#!/usr/bin/env python3
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_test_case_excel(test_cases, output_file="test_cases.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    headers = ["用例ID", "测试场景", "前置条件", "测试步骤", "预期结果", "优先级", "备注"]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_num, test_case in enumerate(test_cases, 2):
        ws.cell(row=row_num, column=1, value=test_case.get("id", "")).border = thin_border
        ws.cell(row=row_num, column=2, value=test_case.get("scenario", "")).border = thin_border
        ws.cell(row=row_num, column=3, value=test_case.get("preconditions", "")).border = thin_border
        ws.cell(row=row_num, column=4, value=test_case.get("steps", "")).border = thin_border
        ws.cell(row=row_num, column=5, value=test_case.get("expected", "")).border = thin_border
        ws.cell(row=row_num, column=6, value=test_case.get("priority", "")).border = thin_border
        ws.cell(row=row_num, column=7, value=test_case.get("notes", "")).border = thin_border

        for col_num in range(1, 8):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    column_widths = [15, 30, 30, 50, 50, 10, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    wb.save(output_file)
    return output_file

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使用方法: python generate_excel.py <json_file> [output_file]")
        sys.exit(1)
    
    json_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else "test_cases.xlsx"
    
    with open(json_file, 'r', encoding='utf-8') as f:
        test_cases = json.load(f)
    
    result_file = create_test_case_excel(test_cases, output_file)
    print(f"Excel 文件已生成: {result_file}")
